"""openpyxl extractor — for .xlsx files.

Reads cell text from all sheets. For each sheet, iterates rows and
cells. Empty cells skipped. No formulas (the cached value is what's
returned, not the formula itself).
"""

from __future__ import annotations

from pathlib import Path

from hermes.memory.extractors import ExtractionResult, logger


def extract(path: Path) -> ExtractionResult:
    """Extract text from a .xlsx file via openpyxl.

    Returns ExtractionResult with text=<cells joined by sheet and row>,
    confidence=1.0, model='openpyxl', error=None on success.

    Text format (one cell per line, with sheet/row context):
        Sheet: <sheetname>
        Row 1: <col1> | <col2> | <col3>
        Row 2: <col1> | <col2>

    This is human-readable AND indexable.

    Error cases:
    - openpyxl not installed: error='openpyxl_not_installed'
    - File not found: error='file_not_found'
    - Corrupted XLSX: error='parse_error: <reason>'
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ExtractionResult(
            text="",
            confidence=1.0,
            model="openpyxl",
            error="openpyxl_not_installed",
        )

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except FileNotFoundError:
        return ExtractionResult(text="", confidence=1.0, model="openpyxl", error="file_not_found")
    except Exception as exc:
        logger.warning(
            "openpyxl_open_error",
            extra={"path": path.as_posix(), "error": str(exc)},
        )
        return ExtractionResult(
            text="",
            confidence=1.0,
            model="openpyxl",
            error=f"parse_error: {exc}",
        )

    parts: list[str] = []
    try:
        for sheet in wb.worksheets:
            parts.append(f"Sheet: {sheet.title}")
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                cells = [str(c) if c is not None else "" for c in row]
                # Drop trailing empty cells
                while cells and not cells[-1]:
                    cells.pop()
                if cells:
                    parts.append(f"Row {row_idx}: " + " | ".join(cells))
    finally:
        wb.close()

    text = "\n".join(parts)
    return ExtractionResult(text=text, confidence=1.0, model="openpyxl", error=None)
